import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


def verificar_dicionario(dicionario):
    tamanho = len(list(dicionario.values())[0])
    for key in dicionario:
        if len(dicionario[key]) != tamanho:
            raise ValueError(f"Todas as listas em '{key}' devem ter o mesmo comprimento.")


def validar_converter_preco(precos):
    return [float(preco.replace(',', '.')) if isinstance(preco, str) else preco for preco in precos]

def ajustar_largura_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def adicionar_formatacao(ws):
    font_bold = Font(bold=True)
    fill_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for cell in ws["1:1"]:
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")

def adicionar_grafico(ws):
    chart = BarChart()
    chart.title = "Comparação de Preços"
    chart.style = 12
    chart.x_axis.title = "Produtos"
    chart.y_axis.title = "Preços (R$)"
    chart.y_axis.majorGridlines = None
    chart.x_axis.titleFont = Font(size=12, bold=True)
    chart.y_axis.titleFont = Font(size=12, bold=True)
    chart.width = 20
    chart.height = 12
    chart.gapWidth = 200  

    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.position = "outEnd" 

    ws.add_chart(chart, "H5")

def gerar_relatorio(cerrano, telentos, valor_disponivel):
    verificar_dicionario(cerrano)
    verificar_dicionario(telentos)

    cerrano['Preço'] = validar_converter_preco(cerrano['Preço'])
    telentos['Preço'] = validar_converter_preco(telentos['Preço'])

    df_cerrano = pd.DataFrame(cerrano)
    df_telentos = pd.DataFrame(telentos)

    df_combined = pd.merge(df_cerrano, df_telentos, on='Produtos', suffixes=('_Cerrano', '_Telentos'))

    df_combined['Mais Barato'] = df_combined.apply(
        lambda row: 'Cerrano' if row['Preço_Cerrano'] < row['Preço_Telentos'] else 'Telentos', axis=1
    )

    def calcula_diferenca(row):
        if row['Mais Barato'] == 'Cerrano':
            diferenca = ((row['Preço_Telentos'] - row['Preço_Cerrano']) / row['Preço_Telentos']) * 100
        else:
            diferenca = ((row['Preço_Cerrano'] - row['Preço_Telentos']) / row['Preço_Cerrano']) * 100
        return round(diferenca, 1)

    df_combined['Diferença (%)'] = df_combined.apply(calcula_diferenca, axis=1)
    df_combined['Diferença Absoluta'] = abs(df_combined['Preço_Cerrano'] - df_combined['Preço_Telentos'])

    total_cerrano = df_combined['Preço_Cerrano'].sum()
    total_telentos = df_combined['Preço_Telentos'].sum()

    economia = total_telentos - total_cerrano if total_cerrano < total_telentos else total_cerrano - total_telentos
    porcentagem_economia = (economia / max(total_cerrano, total_telentos)) * 100

    print(f"Valor disponível para compras: R$ {valor_disponivel:.2f}")
    print(f"Total de compras no Cerrano: R$ {total_cerrano:.2f}")
    print(f"Total de compras no Telentos: R$ {total_telentos:.2f}")
    print(f"Você economizaria R$ {economia:.2f} comprando no mercado mais barato.")
    print(f"Porcentagem de economia: {porcentagem_economia:.1f}%")

  
    arquivo_excel = 'analise_precos.xlsx'
    df_combined.to_excel(arquivo_excel, index=False)

 
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    ajustar_largura_colunas(ws)
    adicionar_formatacao(ws)
    adicionar_grafico(ws)

    ws['H1'] = f"Valor disponível: R$ {valor_disponivel:.2f}"
    ws['H2'] = f"Economia total: R$ {economia:.2f}"
    ws['H3'] = f"Porcentagem de economia: {porcentagem_economia:.1f}%"

    wb.save(arquivo_excel)

    print('Análise concluída e exportada para analise_precos.xlsx com colunas ajustadas e gráficos adicionados.')

valor_disponivel = float(input("Insira o valor disponível para compras: R$ "))


cerrano = {
    'Produtos': ["Oleo", 'Papel higiênico', 'Detergente', 'Sabonete', 'Macarrão', 'Molho', 'Pipoca', 'Salsicha', 'Linguiça', 'Frango', 'Hambúrguer',
                 'Arroz', 'Feijão', 'Café', 'Açúcar', 'Sal', 'Farinha', 'Leite', 'Manteiga', 'Queijo', 'Presunto', 'Pão', 'Biscoito', 'Refrigerante', 'Água'],
    'Preço' : ['5,39', '11,90', '2,59', '3,39', '5,78', '5,56', '5,79', '16,44', '20,06', '27,50', '7,80',
               '18,50', '9,90', '12,30', '3,99', '2,50', '4,90', '4,00', '6,99', '15,00', '9,50', '5,00', '7,50', '8,99', '3,20']
}

telentos = {
    'Produtos': ["Oleo", 'Papel higiênico', 'Detergente', 'Sabonete', 'Macarrão', 'Molho', 'Pipoca', 'Salsicha', 'Linguiça', 'Frango', 'Hambúrguer',
                 'Arroz', 'Feijão', 'Café', 'Açúcar', 'Sal', 'Farinha', 'Leite', 'Manteiga', 'Queijo', 'Presunto', 'Pão', 'Biscoito', 'Refrigerante', 'Água'],
    'Preço':  ['6,50', '13,00', '2,69', '3,80', '10,00', '6,00', '7,00', '18,00', '23,00', '22,00', '9,00',
               '19,00', '10,50', '14,00', '4,50', '2,80', '5,20', '4,50', '7,50', '16,00', '10,00', '5,50', '8,00', '9,50', '3,50'] 
}

gerar_relatorio(cerrano, telentos, valor_disponivel)
